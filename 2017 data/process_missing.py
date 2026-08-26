import pandas as pd
import os
import re
import sys
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
try:
    from krutidev_to_unicode import KrutiDev_to_Unicode
except:
    KrutiDev_to_Unicode = lambda x: x

def clean_kruti_to_unicode(text):
    text = str(text).strip()
    if pd.isna(text) or text == 'nan' or not text:
        return ''
    try:
        # Check if it's Kruti Dev
        if re.search(r'izk|fo0|ua0|d\{k|iV~Vh', text) or (len(re.findall(r'[a-z]', text)) > len(re.findall(r'[A-Z]', text))):
            text = KrutiDev_to_Unicode(text)
        
        # Check if it's Hindi (Devanagari)
        if re.search(r'[\u0900-\u097F]', text):
            text = transliterate(text, sanscript.DEVANAGARI, sanscript.ITRANS)
            text = re.sub(r'([A-Za-z])a(?=\s|$|[^A-Za-z])', r'\1', text)
            text = text.replace('.h', '').replace('aa', 'A').replace('ii', 'I').replace('uu', 'U').upper()
            
    except Exception as e:
        pass
    
    # Final cleanup of names
    text = text.replace(' NOTA ', 'NOTA').replace('NONE OF THE ABOVE', 'NOTA')
    return text.strip()

good_files = [31, 32, 33, 111, 162, 163, 164, 165, 166, 167, 193, 194, 259, 390, 396]

for f in good_files:
    xls = pd.ExcelFile(f'2017_missing/{f}.xls')
    
    # Find best sheet
    best_sheet = None
    max_rows = 0
    for s in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=s, header=None)
            if len(df) > max_rows:
                max_rows = len(df)
                best_sheet = s
        except: pass
        
    df = pd.read_excel(xls, sheet_name=best_sheet, header=None)
    
    # Locate data start
    data_start_row = -1
    booth_col = -1
    for i in range(min(20, len(df))):
        for col in [0, 1, 2]:
            val = str(df.iat[i, col]).strip()
            if val == '1' and str(df.iat[i+1, col]).strip() == '2':
                data_start_row = i
                booth_col = col
                break
        if data_start_row != -1: break
        
    if data_start_row == -1:
        print(f"Failed to find data start in AC {f}")
        continue
        
    cand_row = data_start_row - 1
    # Check if cand_row is all numbers or NaN
    is_numbers = True
    for c in range(booth_col + 1, len(df.columns)):
        val = str(df.iat[cand_row, c]).strip()
        if val != 'nan' and not val.replace('.', '').isdigit():
            is_numbers = False
            break
            
    if is_numbers:
        cand_row = cand_row - 1
        
    if pd.isna(df.iat[cand_row, booth_col + 2]):
        cand_row = cand_row - 1
        
    headers = []
    for col in range(len(df.columns)):
        h = clean_kruti_to_unicode(df.iat[cand_row, col])
        headers.append(h)
        
    # Build extracted dataframe
    extracted = []
    for i in range(data_start_row, len(df)):
        b_val = str(df.iat[i, booth_col]).strip()
        if not b_val.isdigit() or b_val == '0': continue
        
        row_data = {'Booth ID': int(b_val), 'Polling Station Name': f'BOOTH {b_val}', 'Total Electors': 0, 'Male Voters': 0, 'Female Voters': 0, 'Other Voters': 0}
        
        candidates_dict = {}
        total_votes = 0
        nota = 0
        tendered = 0
        
        for col in range(booth_col + 1, len(df.columns)):
            h = headers[col].upper()
            try:
                val = int(float(df.iat[i, col]))
            except:
                val = 0
                
            if 'NOTA' in h or 'NONE' in h:
                nota = val
            elif 'TOTAL' in h or 'KSY' in h or 'YKSX' in h or 'KSYX' in h: # KSY is usually "kul yog" (total) in broken unicode
                total_votes = val
            elif 'TENDER' in h or 'FUFONRR' in h or 'FUFO' in h:
                tendered = val
            elif h and 'VALID' not in h and 'REJECTED' not in h and h != '0' and h != '1' and h != '2':
                # It's a candidate
                candidates_dict[h] = val
                
        # If total_votes wasn't mapped properly, calculate it
        if total_votes == 0:
            total_votes = sum(candidates_dict.values()) + nota
            
        row_data['Total Turnout'] = total_votes
        row_data['Turnout %'] = 0
        row_data['EPIC Voters'] = 0
        row_data['Tendered Voters'] = tendered
        
        for c, v in candidates_dict.items():
            row_data[c] = v
            
        row_data['Total Votes Polled'] = total_votes
        row_data['NOTA'] = nota
        
        cand_vals = list(candidates_dict.values())
        cand_keys = list(candidates_dict.keys())
        
        row_data['Check Sum'] = total_votes - (sum(cand_vals) + nota)
        
        if cand_vals:
            sorted_idx = sorted(range(len(cand_vals)), key=lambda k: cand_vals[k], reverse=True)
            row_data['Winning Candidate'] = cand_keys[sorted_idx[0]]
            row_data['Winning Margin'] = cand_vals[sorted_idx[0]] - cand_vals[sorted_idx[1]] if len(cand_vals) > 1 else cand_vals[sorted_idx[0]]
        else:
            row_data['Winning Candidate'] = ''
            row_data['Winning Margin'] = 0
            
        extracted.append(row_data)
        
    out_df = pd.DataFrame(extracted)
    
    # Save to excel
    output_path = f'excel_outputs/{f}_boothwise_data.xlsx'
    
    header_title = f"UP Assembly General Election 2017 - {f:02d} Detailed Polling Station Data"
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        out_df.to_excel(writer, sheet_name='Booth-wise Data', index=False, startrow=2)
        workbook = writer.book
        worksheet = writer.sheets['Booth-wise Data']
        
        from openpyxl.styles import Font, Alignment
        worksheet.cell(row=1, column=1, value=header_title)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_df.columns))
        worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
    print(f"Successfully wrote {len(out_df)} booths for AC {f} to {output_path}")
