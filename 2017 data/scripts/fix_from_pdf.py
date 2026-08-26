import re
import os
import glob
import pandas as pd
import json

def clean_val(val):
    if not val or val == '-': return '0'
    return val.replace(',', '')

def parse_pdf_text(txt_file, target_acs):
    with open(txt_file, 'r', encoding='utf-8') as f:
        text = f.read()
        
    blocks = re.split(r'={10,}', text)
    data_by_ac = {ac: [] for ac in target_acs}
    
    for block in blocks:
        block = block.strip()
        if not block: continue
        
        # Match AC and Booth
        # "AC 193\n| BOOTH 1" or "AC 385 AJAGARA (S.C.) | BOOTH 26"
        ac_match = re.search(r'AC\s+(\d+).*?\|\s+BOOTH\s+(\d+)', block, re.IGNORECASE | re.DOTALL)
        if not ac_match: continue
        
        ac_no = int(ac_match.group(1))
        if ac_no not in target_acs: continue
        
        booth_no = ac_match.group(2)
        
        # Match Polling Station Name
        ps_match = re.search(r'PS:\s+(.*?)\n', block)
        ps_name = ps_match.group(1).strip() if ps_match else ''
        
        # Match demographics
        demo_match = re.search(r'Electors\s+([\d,]+)\s+\|\s+Voted\s+M\s+([\d,]+)\s+F\s+([\d,]+)\s+O\s+([\d,\-]+)\s+Total\s+([\d,]+)', block)
        if not demo_match: continue
        electors = clean_val(demo_match.group(1))
        male = clean_val(demo_match.group(2))
        female = clean_val(demo_match.group(3))
        other = clean_val(demo_match.group(4))
        turnout = clean_val(demo_match.group(5))
        
        # Match EPIC / Tendered
        epic_match = re.search(r'EPIC-identified\s+([\d,]+)\s+\|\s+Tendered\s+([\d,]+)', block)
        epic = clean_val(epic_match.group(1)) if epic_match else '0'
        tendered = clean_val(epic_match.group(2)) if epic_match else '0'
        
        # Candidates section
        cand_start = block.find('SHARE')
        cand_end = block.find('Valid ')
        if cand_start == -1: cand_start = block.find('VOTES')
        
        cand_text = block[cand_start:cand_end] if cand_start != -1 and cand_end != -1 else ""
        
        cand_lines = cand_text.split('\n')
        candidates_votes = {}
        total_votes_polled = 0
        nota_votes = 0
        
        # Candidate lines look like:
        # 1 ARUN KUMAR MISHRA
        # BSP
        # 6
        # 0.71%
        
        # We can extract numbers that look like votes. Wait, it's spread over 4 lines per candidate!
        # Let's extract lines that start with a number followed by a name
        cand_names = re.findall(r'^(\d+)\s+(.*?)$', cand_text, re.MULTILINE)
        
        # To robustly extract, let's just find the pattern:
        # \n(\d+)\s+([A-Za-z\s.]+)\n(.*?)\n([\d,]+)\n
        
        # Actually, split cand_text by numbers at the start of the line
        chunks = re.split(r'\n(?=\d+\s+[A-Za-z])', '\n' + cand_text.strip())
        for chunk in chunks:
            chunk = chunk.strip()
            if not chunk: continue
            
            # format:
            # 1 ARUN KUMAR MISHRA\nBSP\n6\n0.71%
            lines = [l.strip() for l in chunk.split('\n') if l.strip()]
            if len(lines) >= 3:
                name_match = re.match(r'^\d+\s+(.*)$', lines[0])
                if not name_match: continue
                name = name_match.group(1).strip()
                
                # The vote is usually the second to last line if percentage is last
                # Or we can just find the first purely numeric line from the bottom
                vote = '0'
                for l in reversed(lines):
                    if '%' not in l and re.match(r'^[\d,]+$', l):
                        vote = clean_val(l)
                        break
                
                if 'NOTA' in name.upper():
                    nota_votes = int(vote)
                else:
                    candidates_votes[name] = int(vote)
                    total_votes_polled += int(vote)
                    
        total_votes_polled += nota_votes
        
        row_dict = {
            'Booth ID': int(booth_no),
            'Polling Station Name': ps_name,
            'Total Electors': int(electors),
            'Male Voters': int(male),
            'Female Voters': int(female),
            'Other Voters': int(other),
            'Total Turnout': int(turnout),
            'Turnout %': int(turnout) / int(electors) if int(electors) > 0 else 0,
            'EPIC Voters': int(epic),
            'Tendered Voters': int(tendered)
        }
        row_dict.update(candidates_votes)
        row_dict['Total Votes Polled'] = total_votes_polled
        row_dict['NOTA'] = nota_votes
        
        # Checksum calculation: Total Votes Polled - (Sum of candidate votes + NOTA)
        row_dict['Check Sum'] = total_votes_polled - (sum(candidates_votes.values()) + nota_votes)
        
        data_by_ac[ac_no].append(row_dict)
        
    return data_by_ac

def save_excel(data, ac_no):
    if not data:
        print(f"No data found for AC {ac_no}")
        return
        
    df = pd.DataFrame(data)
    df.sort_values(by='Booth ID', inplace=True)
    
    # Calculate Winners
    cand_cols = [c for c in df.columns if c not in ['Booth ID', 'Polling Station Name', 'Total Electors', 'Male Voters', 'Female Voters', 'Other Voters', 'Total Turnout', 'Turnout %', 'EPIC Voters', 'Tendered Voters', 'Total Votes Polled', 'NOTA', 'Check Sum']]
    
    def get_winner(row):
        votes = row[cand_cols].fillna(0)
        if votes.empty or votes.sum() == 0: return ""
        return votes.idxmax()
        
    def get_margin(row):
        votes = row[cand_cols].fillna(0).sort_values(ascending=False)
        if len(votes) >= 2: return votes.iloc[0] - votes.iloc[1]
        return 0
        
    df['Winning Candidate'] = df.apply(get_winner, axis=1)
    df['Winning Margin'] = df.apply(get_margin, axis=1)
    
    # Reorder columns
    col_order = ['Booth ID', 'Polling Station Name', 'Total Electors', 'Male Voters', 'Female Voters', 'Other Voters', 'Total Turnout', 'Turnout %', 'EPIC Voters', 'Tendered Voters'] + cand_cols + ['Total Votes Polled', 'NOTA', 'Check Sum', 'Winning Candidate', 'Winning Margin']
    df = df[[c for c in col_order if c in df.columns]]
    
    # Header format
    constituency_name = ""
    try:
        with open('constituencies.json', 'r') as cf:
            cmap = json.load(cf)
            constituency_name = cmap.get(str(ac_no), "")
    except:
        pass
        
    header_title = f"UP Assembly General Election 2017 - {ac_no:02d} Detailed Polling Station Data"
    output_path = f"excel_outputs/{ac_no}_boothwise_data.xlsx"
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Booth-wise Data', index=False, startrow=2)
        
        workbook = writer.book
        worksheet = writer.sheets['Booth-wise Data']
        
        worksheet.cell(row=1, column=1, value=header_title)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        
        from openpyxl.styles import Font, Alignment
        worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        if constituency_name:
            worksheet.cell(row=2, column=1, value=f"- {constituency_name} Cons.")
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
            worksheet.cell(row=2, column=1).font = Font(bold=True, size=12)
            worksheet.cell(row=2, column=1).alignment = Alignment(horizontal='center')
            
    print(f"Successfully saved {output_path} with {len(df)} booths.")

if __name__ == "__main__":
    # 1. Farrukhabad (193, 194)
    print("Processing Farrukhabad...")
    far_data = parse_pdf_text('Farrukhabad_nolayout.txt', [193, 194])
    save_excel(far_data[193], 193)
    save_excel(far_data[194], 194)
    
    # 2. Varanasi (390)
    print("Processing Varanasi...")
    var_data = parse_pdf_text('Varanasi_nolayout.txt', [390])
    save_excel(var_data[390], 390)
    
    # 3. Mirzapur (396)
    print("Processing Mirzapur...")
    mir_data = parse_pdf_text('Mirzapur_nolayout.txt', [396])
    save_excel(mir_data[396], 396)
