import os
import glob
import re
import pandas as pd
from collections import defaultdict
import xml.etree.ElementTree as ET
import warnings
import krutidev_to_unicode
from krutidev_to_unicode import KrutiDev_to_Unicode
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

warnings.simplefilter(action='ignore', category=FutureWarning)

def clean_xml_content(content):
    content = re.sub(r'<\?xml[^>]*\?>', '', content)
    return f'<root>{content}</root>'

def extract_candidates(root):
    max_base_x = 0
    for cell in root.findall('.//cell'):
        if cell.get('p') != '1': continue
        try:
            x = int(cell.get('x', 0))
            text = ''.join(cell.itertext()).lower()
            text_nospace = text.replace(' ', '')
            if any(kw in text_nospace for kw in ['turnout', 'epic', 'tendered', 'ten-dere', 'ten-dered']):
                if x > max_base_x: max_base_x = x
        except: pass

    cands_x_groups = defaultdict(list)
    for cell in root.findall('.//cell'):
        if cell.get('p') != '1': continue
        try:
            x, w, h = int(cell.get('x', 0)), int(cell.get('w', 0)), int(cell.get('h', 0))
            if x > max_base_x + 30:
                cands_x_groups[(w, h)].append(cell)
        except: pass

    best_row_key = None
    max_alpha_cells = 0
    
    for k, cells in cands_x_groups.items():
        alpha_count = sum(1 for c in cells if any(char.isalpha() for char in ''.join(c.itertext())))
        text_concat = ' '.join(''.join(c.itertext()).lower() for c in cells)
        if 'votes secured' in text_concat or 'party affiliation' in text_concat or 's.no.' in text_concat.replace(' ',''):
            continue
            
        if alpha_count > max_alpha_cells:
            max_alpha_cells = alpha_count
            best_row_key = k

    if not best_row_key: return [], [], max_base_x

    cells = sorted(cands_x_groups[best_row_key], key=lambda c: int(c.get('x', 0)))
    candidates = [''.join(c.itertext()).strip().replace('\n', ' ') for c in cells]
    cand_xs = [int(c.get('x', 0)) for c in cells]
    
    return candidates, cand_xs, max_base_x

def is_valid_data_row(row, base_name):
    num_numeric = sum(1 for x in row if str(x).replace(',', '').replace('-', '').isdigit())
    if num_numeric < 5: return False
    
    # Explicitly reject numbering rows (e.g. 1, 2, 3, 4, 5...)
    if len(row) >= 5 and row[1:5] == ['2', '3', '4', '5']:
        return False
        
    if str(row[0]).lstrip('0') == str(base_name).lstrip('0') and len(row) > 1 and str(row[1]).isdigit(): return True
    if str(row[0]).isdigit() and int(row[0]) > 0: return True
    if len(row) > 1 and str(row[1]).isdigit() and int(row[1]) > 0: return True
    return False

def parse_blob(content, base_name):
    text = re.sub(r'<[^>]+>', ' ', content)
    text = re.sub(r'\s+', ' ', text)
    
    segments = re.split(fr'(?=\b{base_name}\s+\d+\s+[A-Za-z])', text)
    rows = []
    for seg in segments:
        seg = seg.strip()
        if not seg.startswith(base_name): continue
        match = re.match(fr'^{base_name}\s+(\d+)\s+(.+?)\s+([\d,]+|-)\s+([\d,]+|-)\s+([\d,]+|-)\s+([\d,]+|-)\s+([\d,]+|-)', seg)
        if match:
            booth_no = match.group(1)
            name = match.group(2)
            electors = match.group(3).replace(',', '')
            male = match.group(4).replace(',', '')
            female = match.group(5).replace(',', '')
            other = match.group(6).replace(',', '')
            turnout = match.group(7).replace(',', '')
            
            votes_match = re.findall(r'\b\d+\s+[A-Za-z.\-]+\s+(\d+|-)', seg[match.end():])
            row = [base_name, booth_no, name, electors, male, female, other, turnout]
            for v in votes_match:
                row.append("0" if v == "-" else v)
            rows.append(row)
    return pd.DataFrame(rows)

def parse_xml_to_dataframe(xml_file):
    with open(xml_file, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    
    base_name = os.path.basename(xml_file).replace('.xml', '')
    
    try:
        root = ET.fromstring(clean_xml_content(content))
    except Exception as e:
        print(f"Error parsing {xml_file}: {e}")
        return None, []

    candidates, cand_xs, max_base_x = extract_candidates(root)
    
    pages = defaultdict(list)
    for cell in root.findall('.//cell'):
        try:
            h, p, w, x, y = int(cell.get('h', 0)), int(cell.get('p', 0)), int(cell.get('w', 0)), int(cell.get('x', 0)), int(cell.get('y', 0))
            text = ''.join(cell.itertext()).strip().replace('\n', ' ')
            if text:
                pages[p].append({'h': h, 'w': w, 'x': x, 'y': y, 'text': text})
        except: pass

    all_rows = []
    for p in sorted(pages.keys()):
        cells = pages[p]
        row_groups = defaultdict(list)
        for cell in cells:
            row_groups[(cell['w'], cell['h'])].append(cell)
            
        sorted_keys = sorted(row_groups.keys(), key=lambda k: k[0])
        for key in sorted_keys:
            row_cells = row_groups[key]
            row_cells.sort(key=lambda item: item['x'])
            
            if not cand_xs:
                all_rows.append([c['text'] for c in row_cells])
                continue
                
            base_texts = [c['text'].strip() for c in row_cells if c['x'] <= max_base_x + 30]
            cand_cells = [c for c in row_cells if c['x'] > max_base_x + 30]
            
            if not base_texts or not cand_cells: continue
            
            total_votes_cell = cand_cells.pop()
            total_votes_text = total_votes_cell['text'].strip()
                
            candidate_votes = []
            for i in range(len(cand_xs)):
                x_start = cand_xs[i] - 30
                if i < len(cand_xs) - 1:
                    x_end = cand_xs[i+1] - 30
                else:
                    x_end = float('inf')
                    
                c_cells = [c for c in cand_cells if x_start <= c['x'] < x_end]
                
                vote = '0'
                for c in reversed(c_cells):
                    t = c['text'].strip()
                    if t == '-':
                        vote = '0'
                        break
                    if str(t).replace(',', '').replace('-', '').isdigit():
                        vote = t
                        break
                candidate_votes.append(vote)
                
            clean_row = [t for t in base_texts if t] + candidate_votes + [total_votes_text]
            all_rows.append(clean_row)

    data_rows = []
    for row in all_rows:
        if len(row) > 5:
            clean_row = []
            for item in row:
                if ' ' in item and len(item.split()) > 3 and not any(c.isalpha() for c in item):
                    clean_row.extend(item.split())
                else:
                    clean_row.append(item)
                    
            if is_valid_data_row(clean_row, base_name):
                # Unmerge Booth ID and Name if they got grouped into clean_row[0]
                match = re.match(r'^(\d+)\s*([^\d].*)$', str(clean_row[0]))
                if match and len(clean_row) > 1 and str(clean_row[1]).replace(',', '').isdigit():
                    clean_row.insert(0, match.group(1))
                    clean_row[1] = match.group(2)
                
                if str(clean_row[0]).lstrip('0') != str(base_name).lstrip('0') or not str(clean_row[1]).replace(',', '').isdigit():
                    clean_row.insert(0, base_name)
                data_rows.append(clean_row)

    if not data_rows:
        df_blob = parse_blob(content, base_name)
        return df_blob, candidates
        
    return pd.DataFrame(data_rows), candidates

def format_exact_excel(df, candidates, output_path, base_name):
    try:
        if df.empty: return False
        df = df.copy()
        
        def is_shifted(df_col):
            # Check if the column is mostly non-numeric (e.g. AC Name text)
            numeric_col = pd.to_numeric(df_col, errors='coerce')
            return numeric_col.isna().mean() > 0.5
        
        # 30 files have an extra AC Name column pushing Booth ID to df[2]
        shift = 1 if is_shifted(df[1]) else 0

        # Clean numeric columns (avoiding the text column, which is index 2 + shift)
        for col in df.columns:
            if col != 2 + shift:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '').str.replace('-', '0'), errors='coerce')
        
        def clean_polling_station_name(name):
            name = str(name).strip()
            if not name: return name
            
            # Remove leading numbers (which are just the booth serial number accidentally prepended)
            name = re.sub(r'^\d+[\s.-]*', '', name)
            
            # Step 1: Convert Kruti Dev to Hindi Unicode if needed
            if re.search(r'izk|fo0|ua0|d\{k|iV~Vh', name) or (len(re.findall(r'[a-z]', name)) > len(re.findall(r'[A-Z]', name))):
                name = KrutiDev_to_Unicode(name)
            # Step 2: Transliterate Hindi (Devanagari) to English if present
            if re.search(r'[\u0900-\u097F]', name):
                name = transliterate(name, sanscript.DEVANAGARI, sanscript.ITRANS)
                # Remove trailing schwa 'a' added by ITRANS
                name = re.sub(r'([A-Za-z])a(?=\s|$|[^A-Za-z])', r'\1', name)
                name = name.replace('.h', '')
                name = name.replace('aa', 'A').replace('ii', 'I').replace('uu', 'U')
                name = name.upper()
            return name

        base_cols = {
            'Booth ID': df[1 + shift],
            'Polling Station Name': df[2 + shift].apply(clean_polling_station_name),
            'Total Electors': df[3 + shift],
            'Male Voters': df[4 + shift],
            'Female Voters': df[5 + shift],
            'Other Voters': df[6 + shift],
            'Total Turnout': df[7 + shift]
        }
        
        out_df = pd.DataFrame(base_cols)
        
        # Remove the column numbering row (which typically has 3,4,5 or 4,5,6 in the numeric columns)
        out_df = out_df[~((out_df['Total Electors'].astype(str).str.strip().isin(['3', '4', '3.0', '4.0'])) & 
                          (out_df['Male Voters'].astype(str).str.strip().isin(['4', '5', '4.0', '5.0'])) &
                          (out_df['Female Voters'].astype(str).str.strip().isin(['5', '6', '5.0', '6.0'])))]
        
        out_df['Turnout %'] = out_df['Total Turnout'] / out_df['Total Electors']
        
        # We know clean_row has format: base_cols + candidate_votes + total_votes
        # So base_cols count = df.shape[1] - (len(candidates) + 1)
        # But candidates might include NOTA which is handled, so num_cands is len(candidates)
        # Actually, candidates length is the exact number of vote columns appended.
        num_vote_cols = len(candidates) + 1
        num_base_cols = df.shape[1] - num_vote_cols
        
        # Base columns always include up to Total Turnout (index 7 + shift).
        # That requires 8 + shift columns (0 to 7+shift).
        if num_base_cols > 8 + shift:
            out_df['EPIC Voters'] = df[8 + shift]
            if num_base_cols > 9 + shift:
                out_df['Tendered Voters'] = df[9 + shift]
            else:
                out_df['Tendered Voters'] = 0
        else:
            out_df['EPIC Voters'] = 0
            out_df['Tendered Voters'] = 0
            
        data_start = num_base_cols
        candidate_data = df.iloc[:, data_start:]
        num_candidates = min(len(candidates), candidate_data.shape[1] - 1)
        
        for i in range(num_candidates):
            col_name = candidates[i]
            if 'NOTA' in col_name.upper():
                col_name = 'NOTA'
            out_df[col_name] = candidate_data.iloc[:, i]
            
        if candidate_data.shape[1] > num_candidates:
            out_df['Total Votes Polled'] = candidate_data.iloc[:, num_candidates]
        else:
            out_df['Total Votes Polled'] = candidate_data.sum(axis=1)
            
        if 'NOTA' not in out_df.columns:
            out_df['NOTA'] = 0
            
        candidate_cols = [c for c in out_df.columns if c not in ['Booth ID', 'Polling Station Name', 'Total Electors', 'Male Voters', 'Female Voters', 'Other Voters', 'Total Turnout', 'Turnout %', 'EPIC Voters', 'Tendered Voters', 'NOTA', 'Total Votes Polled', 'Check Sum', 'Winning Candidate', 'Winning Margin']]
        
        out_df['Check Sum'] = out_df['Total Votes Polled'] - (out_df[candidate_cols].sum(axis=1) + out_df['NOTA'].fillna(0))
        
        def get_winner(row):
            votes = row[candidate_cols].fillna(0)
            if votes.empty or votes.sum() == 0: return ""
            return votes.idxmax()
            
        def get_margin(row):
            votes = row[candidate_cols].fillna(0).sort_values(ascending=False)
            if len(votes) >= 2: return votes.iloc[0] - votes.iloc[1]
            return 0
            
        out_df['Winning Candidate'] = out_df.apply(get_winner, axis=1)
        out_df['Winning Margin'] = out_df.apply(get_margin, axis=1)

        constituency_name = ""
        import json
        import os
        try:
            with open('constituencies.json', 'r') as cf:
                cmap = json.load(cf)
                constituency_name = cmap.get(str(base_name), "")
        except:
            pass

        try:
            bn_int = int(base_name)
            header_title = f"UP Assembly General Election 2017 - {bn_int:02d} Detailed Polling Station Data"
        except:
            header_title = f"UP Assembly General Election 2017 - {base_name} Detailed Polling Station Data"

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            if constituency_name:
                out_df.to_excel(writer, sheet_name='Booth-wise Data', index=False, startrow=2)
            else:
                out_df.to_excel(writer, sheet_name='Booth-wise Data', index=False, startrow=1)
                
            workbook = writer.book
            worksheet = writer.sheets['Booth-wise Data']
            
            # Format title row
            worksheet.cell(row=1, column=1, value=header_title)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_df.columns))
            
            from openpyxl.styles import Font, Alignment
            title_font = Font(bold=True, size=14)
            subtitle_font = Font(bold=True, size=12)
            
            worksheet.cell(row=1, column=1).font = title_font
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            
            # Format subtitle row
            if constituency_name:
                worksheet.cell(row=2, column=1, value=f"- {constituency_name} Cons.")
                worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(out_df.columns))
                worksheet.cell(row=2, column=1).font = subtitle_font
                worksheet.cell(row=2, column=1).alignment = Alignment(horizontal='center')
            
        return True
    except Exception as e:
        print(f"Error formatting {output_path}: {e}")
        return False

if __name__ == "__main__":
    input_dir = r"c:\Users\ASUS\Desktop\freelancing\political analysis\2017 data\upvidhansabha2017"
    output_dir = r"c:\Users\ASUS\Desktop\freelancing\political analysis\2017 data\excel_outputs"
    os.makedirs(output_dir, exist_ok=True)
    
    xml_files = glob.glob(os.path.join(input_dir, "*.xml"))
    
    success = 0
    failed = 0
    
    print(f"Processing {len(xml_files)} files with accurate format...")
    for file in xml_files:
        base_name = os.path.basename(file).replace('.xml', '')
        output_path = os.path.join(output_dir, f"{base_name}_boothwise_data.xlsx")
        
        df, candidates = parse_xml_to_dataframe(file)
        if df is not None and not df.empty:
            if format_exact_excel(df, candidates, output_path, base_name):
                success += 1
            else:
                failed += 1
        else:
            print(f"Could not extract data from {base_name}.xml")
            failed += 1
            
    print(f"Finished. Success: {success}, Failed: {failed}")
